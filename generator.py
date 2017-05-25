# Copyright (c) 2017 Niayesh Ilkhani

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#    http://www.apache.org/licenses/LICENSE-2.0

# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

"""Submeter Bill Generator

This module is designed to take an Excel sheet of tenant data, as well
as a utility bill template, and produce bills for every tenant for the
selected billing period.
"""

from os import environ, makedirs, listdir, rmdir
from os.path import exists, abspath
from subprocess import run

from collections import OrderedDict
from decimal import Decimal, getcontext, ROUND_HALF_EVEN
from datetime import datetime
from textwrap import fill, wrap

from tkinter import *
from tkinter import filedialog, messagebox
from tkinter.ttk import *

from openpyxl import load_workbook
from docxtpl import DocxTemplate, InlineImage

from matplotlib import use, rcParams
use("Agg")
from matplotlib import pyplot


GAL_TO_M3 = Decimal('0.00378541')  # as on bill template (correct to 6 sig figs)

charts_dir = "Charts"
bills_dir = "Bills"

data_sheet = "DataNewest"

environ.setdefault("PYPANDOC_PANDOC", "C:/Program Files (x86)/Pandoc")
getcontext().rounding = ROUND_HALF_EVEN  # so it's not locale-specific
# Matplotlib settings
rcParams.update({'xtick.labelsize': 'small',
                 'figure.dpi': 150.0,
                 'figure.figsize': [8/3, 2.0]})


class InvalidDataError(Exception):
    def __init__(self, message, tk: Tk):
        tk.quit()
        super().__init__(message)


class Tenant:
    def __init__(self, unit_no, document, context):
        self.unit_no = unit_no
        self.document = document
        self.context = context.copy()  # shallow copy is enough (no mutable values)
        self.chart_path = "{}/unit_{}_history.png".format(charts_dir, self.unit_no)
        self.bill_path = "{}/unit_{}_{}_bill.docx".format(
            bills_dir, self.unit_no, bills_dir[6:])  # formatted period-name

        self.four_recent_bills = []  # type: [Decimal]

        self.context['CRead'] = self.context['PRead'] = None

        # prev_balance left at 0 for now, can be used later
        self.context['PrevBalance'] = Decimal(0.00)

    def get_addr_info(self, row):
        """
        Fills in this tenant's name and address information from
        the given row from the TenantInfo sheet of the Excel document.
        Returns True on success.

        :param List[Cell] row: this tenant's row in the TenantInfo sheet
        :return: bool
        """
        if row[2].value == "Vacant":  # no actual tenant
            return False
        self.context['Name'] = row[2].value
        self.context['MeterNo'] = str(row[1].value)
        self.context['AccountNo'] = '-'  # not used right now, but needed for template
        self.context['ServiceAddr'] = row[3].value
        self.context['BillingAddr'] = row[7].value
        self.context['BillingCity'] = row[8].value
        self.context['BillingProv'] = row[9].value
        self.context['BillingPostal'] = row[10].value
        return True  # all is well

    def calculate_bills(self, row, index):
        """
        Calculates the current bill and up to 3 previous bills
        for this tenant. Returns True on success.

        :param List[Cell] row: the Excel row for this tenant
        :param int index: the index to the current period to calculate
        :return: bool
        """
        # Need unit conversion:
        curr = row[index].value
        prev = row[index - 1].value
        if (isinstance(curr, str) or curr <= 0 or
                isinstance(prev, str) or prev < 0):  # zero-value prev is ok
            return False
        self.context['CRead'] = _stround(Decimal(curr) * GAL_TO_M3)
        self.context['PRead'] = _stround(Decimal(prev) * GAL_TO_M3)
        self.four_recent_bills.append(Decimal(self.context['CRead']) -
                                      Decimal(self.context['PRead']))
        if self.four_recent_bills[0] == 0:  # no consumption this period
            return False
        elif self.four_recent_bills[0] < 0:
            print("Something is very wrong with unit", row[0].value)
            return False
        self.context['Cons'] = Decimal(self.four_recent_bills[0] *
                                       self.context['Rate'])
        self.context['TotalDue'] = _stround(self.context['PrevBalance'] +
                                            self.context['Cons'])
        # Normalise to strings
        self.context['Cons'] = _stround(self.context['Cons'])
        self.context['PrevBalance'] = _stround(self.context['PrevBalance'])
        self.context['Rate'] = str(round(self.context['Rate'], 4))  # different rounding than _stround
        self.context['AmCons'] = _stround(self.four_recent_bills[0])

        count = 0
        index -= 1
        # print(index)
        while count < 3 and index > 2:
            index -= 1
            curr = prev
            prev = row[index].value
            # print(index, curr, prev)
            if isinstance(prev, str) or prev < 0:  # zero values are ok
                break  # don't try to add any more columns to the graph
            # Still need unit conversion:
            self.four_recent_bills.append(
                GAL_TO_M3 * (Decimal(curr) - Decimal(prev)))
            count += 1
        while count < 3:  # fill out the 4 column spaces
            self.four_recent_bills.append(Decimal(0))
            count += 1
        # print(row[0].value, ":", *self.four_recent_bills, sep=", ")
        return True  # all is well

    def generate_chart(self, index, periods):
        """

        :type index: int
        :type periods: OrderedDict
        :return: None
        """

        def autolabel(rects, axes):
            """
            Attach a text label above each bar displaying its height
            Thanks to the API and Lindsey Kuper at composition.al.
            """
            (y_bottom, y_top) = axes.get_ylim()
            y_height = y_top - y_bottom
            for rect in rects:
                height = rect.get_height()
                if height > 0:
                    label_position = height + (y_height * 0.05)
                    axes.text(rect.get_x() + rect.get_width() / 2.,
                              label_position, str(round(height, 2)),
                              ha='center', va='bottom')

        bills = self.four_recent_bills[::-1]  # reverse
        rng = range(4)
        fig, ax = pyplot.subplots()  # type: Figure, Axes

        ax.set_xticks(rng)
        labels = []
        i = -2
        periods = list(periods.items())
        while i >= -5:
            text = fill(periods[index + i][0], 10) if index + i >= 0 else ''
            labels.insert(0, text)
            i -= 1
        ax.set_xticklabels(labels)
        ax.set_xticklabels(ax.xaxis.get_majorticklabels(), rotation=90)

        ax.axes.get_yaxis().set_visible(False)
        ax.spines['top'].set_visible(False)
        ax.spines['left'].set_visible(False)
        ax.spines['right'].set_visible(False)

        chart = ax.bar(rng, bills, 1/1.5, color="blue")
        chart[3].set_color("cyan")
        autolabel(chart, ax)
        fig.subplots_adjust(bottom=0.4)
        fig.savefig(self.chart_path)
        pyplot.close(fig)

    def generate_bill(self):
        self.context['Chart'] = InlineImage(self.document, self.chart_path)
        self.document.render(self.context)
        self.document.save(self.bill_path)


class BillGenerator(Tk):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.title("Submeter Bill Generator")

        self.template = StringVar(self)
        self.periods = self.period_index = self.data = self.contacts = None
        self.tenants = []
        self.context = {}

        self.run()

    def run(self):
        period_menu = None
        option = StringVar(self)
        datafile = StringVar(self)

        def get_data_file(*args):
            fn = filedialog.askopenfilename(
                filetypes=[("Excel files", ".xlsx")], parent=self)
            if not fn:
                messagebox.showerror("Opening File",
                                     "Cannot open file. Please try again.",
                                     parent=self)
            else:
                datafile.set(fn)

        def read_data_file(*args):
            ok_button.pack_forget()
            message.configure(text="Data file open and reading.\n"
                                   "Please wait.")
            self.update_idletasks()
            wb = load_workbook(filename=datafile.get(), data_only=True)
            self.data = wb[data_sheet]
            self.contacts = wb["TenantInfo"]
            self.periods = OrderedDict((c.value, i + 2) for (i, c)
                                       in enumerate(self.data[1][2:])
                                       if c.value is not None)

            # Choose from available billing periods
            nonlocal period_menu
            last = self.periods.popitem()
            self.periods[last[0]] = last[1]
            last = last[0]
            period_menu = OptionMenu(mainframe, option, last,
                                     *self.periods.keys())
            message.configure(text="Please select a billing period_index\n"
                                   "to produce bills for:")
            ok_button.configure(text="OK", command=select_period)
            period_menu.grid(row=2, column=1)
            ok_button.pack(pady=10, padx=10)

        def select_period(*args):
            nonlocal option
            self.period_index = self.periods[option.get()]
            if any(len(x) > 10 for x in wrap(option.get(), 10)):
                messagebox.showwarning("Format Warning",
                                       "Warning: This period name '%s' "
                                       "contains word(s) \nlonger than 10 "
                                       "characters. This may result in a "
                                       "formatting error.\nChange the period "
                                       "name in the Excel file in accordance "
                                       "with \nthe naming standard, or proceed "
                                       "with caution." % option.get(),
                                       parent=self)
            # Check that there is at least one reading for this time period
            readings = self.data[chr(ord('A') + self.period_index)][6:]
            # for r in readings: print(r.value, type(r.value))
            if all(isinstance(r.value, str) for r in readings):  # all errors
                raise InvalidDataError("The selected billing period, " +
                                       option.get() + ", does not "
                                       "contain any valid readings.", self)

            # Add billing period to Charts and Bills folder names
            global charts_dir, bills_dir
            option = option.get().replace('- ', '-').replace(
                ' -', '-').replace(' ', '_')  # turn "X Y - Z" into "X_Y-Z"
            charts_dir += '/' + option
            bills_dir += '/' + option

            period_menu.grid_forget()
            message.configure(text="Please select a Word file to use\n"
                                   "as a template for the bills:")
            self.template.trace('w', use_template_file)
            ok_button.configure(text="Choose File", command=get_template_file)

        def get_template_file(*args):
            fn = filedialog.askopenfilename(
                filetypes=[("Word files", ".docx")], parent=self)
            if not fn:
                messagebox.showerror("Opening File",
                                     "Cannot open file. Please try again.",
                                     parent=self)
            else:
                self.template.set(fn)

        def use_template_file(*args):
            ok_button.pack_forget()
            message.configure(text="Template file opened. Generating bills.\n"
                                   "Please wait, as this may take some time.")
            self.update_idletasks()
            self.template = self.template.get()

            # Do the actual generation
            get_period_info()
            initialise_tenants()
            generate_charts()
            generate_bills()
            generate_pdfs()

        def get_period_info():
            period = self.data[chr(ord('A') + self.period_index)]
            # Already datetime objects (thanks, openpyxl)
            # self.start_date = datetime.strptime(period[1].value, "%Y-%m-%d")
            # self.end_date = datetime.strptime(period[2].value, "%Y-%m-%d")
            self.context['StartDate'] = period[1].value.strftime("%b %d/%y")  # Jan 31/17
            self.context['EndDate'] = period[2].value.strftime("%b %d/%y")
            self.context['NumDays'] = str(period[3].value)
            self.context['Rate'] = Decimal(period[4].value)
            self.context['DueDate'] = period[5].value.strftime("%b %d/%y")

        def initialise_tenants():
            for i, unit_row in enumerate(self.data.iter_rows(min_row=7)):
                unit = unit_row[0].value
                if unit is None:  # no unit name
                    continue
                if unit_row[self.period_index] is None or \
                        isinstance(unit_row[self.period_index].value, str):
                    print("Skipping tenant (invalid reading):", unit)
                    continue  # no value or invalid value
                t = Tenant(unit_no=str(unit),
                           document=DocxTemplate(self.template),
                           context=self.context)
                contact_row = self.contacts[i + 2]
                assert contact_row[0].value == unit, \
                    "Mismatch in the tenant order of {} and " \
                    "TenantInfo sheets ({} != {})".format(
                        data_sheet, contact_row[0].value, unit)
                if not t.get_addr_info(contact_row):
                    del t  # vacant
                    print("Skipping tenant (vacant):", unit)
                    continue
                if not t.calculate_bills(unit_row, self.period_index):
                    del t  # no usage
                    print("Skipping tenant (no usage):", unit)
                    continue

                self.tenants.append(t)  # all is well

        def generate_charts():
            assert not exists(charts_dir), \
                "'{}' folder already exists.".format(charts_dir)
            makedirs(charts_dir)
            for tenant in self.tenants:
                tenant.generate_chart(self.period_index, self.periods)

        def generate_bills():
            assert not exists(bills_dir), \
                "'{}' folder already exists.".format(bills_dir)
            makedirs(bills_dir)
            for tenant in self.tenants:
                tenant.generate_bill()

        def generate_pdfs():
            # ok_button.pack_forget()
            # exit_button.pack_forget()
            # message.configure(text="Converting to PDFs. Please wait.")
            # self.update_idletasks()
            for tenant in self.tenants:
                run(["wscript", "doc2pdf.vbs", tenant.bill_path])
            # message.configure(text="Conversion complete! Press Exit to close.")
            # exit_button.pack(padx=10, pady=10)
            almost_close()

        def almost_close():
            path = abspath(bills_dir)
            message.configure(text="Bill generation complete!\n"
                                   "Bills have been stored in \n%s.\n"
                                   "Press Open Folder to open this folder \n"
                                   "in Explorer. Press Exit to close." % path)
            ok_button.configure(text="Open Folder",
                                command=lambda: run("explorer " + path))
            exit_button.configure(command=lambda: self.quit())
            ok_button.pack(padx=10, pady=5)
            exit_button.pack(padx=10, pady=5)

        # Add a grid
        mainframe = Frame(self)
        mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
        mainframe.columnconfigure(0, weight=1)
        mainframe.rowconfigure(0, weight=1)
        mainframe.pack(padx=100, pady=50)

        # Select data file
        message = Label(mainframe, text="Select a data file:", justify=CENTER)
        message.grid(row=1, column=1)
        exit_button = Button(self, text="Exit")
        ok_button = Button(self, text="Choose File", command=get_data_file)
        datafile.trace('w', read_data_file)
        ok_button.pack(padx=10, pady=10)


def _stround(num):
    """
    Round num to two decimal places and return as a string.

    :type num: Decimal
    :return: str
    """
    assert isinstance(num, Decimal), "Not a Decimal type"
    return str(round(Decimal(num), 2))


if __name__ == "__main__":
    try:
        root = BillGenerator()
        root.mainloop()
    except Exception as e:
        if exists(charts_dir) and not listdir(charts_dir):
            rmdir(charts_dir)
        if exists(bills_dir) and not listdir(bills_dir):
            rmdir(bills_dir)
        raise e
